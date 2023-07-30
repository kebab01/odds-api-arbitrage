import json
import openpyxl
from models import *
import os
from dotenv import set_key

def cartesian_product(lst:list[Bookmaker], n:int):
    #Get the cartesion product of bookies
    if n == 1:
        return [(x,) for x in lst]
    return [(x,) + t for t in cartesian_product(lst, n - 1) for x in lst]


def load_comps():
    with open('sports.json','r') as f:
        comps=[i for i in json.load(fp=f)]
        valid_comps=list(filter(lambda x: x['active']==True, comps))
        return [i['key'] for i in valid_comps]
    
def prepare_data(sports:list[BetObject]):
    
    processed=[]
    for i in sports:
        tmp=[
            i.sport_title,
            i.commence_time, 
            i.home_team, 
            i.actions[0].action.price,
            i.actions[0].agency,
            i.away_team, 
            i.actions[1].action.price,
            i.actions[1].agency,
        ]
        try:
            # Depends if sport can have a draw
            tmp.append(i.actions[2].action.price)
            tmp.append(i.actions[2].agency)
        except IndexError:
            tmp.append(0)
            tmp.append(0)

        tmp.append(i.total_odds)
        processed.append(tmp)

    return processed


def write_to_excel(sports:list[BetObject]):

    sports=prepare_data(sports)
    
    wb=openpyxl.load_workbook(filename="bets.xlsx")
    sheet=wb['raw']

    sheet.delete_rows(2, sheet.max_row)

    for i, row in enumerate(sports):
        for x, col in enumerate(row):
            sheet.cell(column=x+1, row=i+2, value=row[x])
    
    wb.save("bets.xlsx")

def convert_frm_iso(iso_string) -> str:

    # Convert ISO 8601 string to datetime object
    dt_object = datetime.strptime(iso_string, "%Y-%m-%dT%H:%M:%SZ")
    formatted_date = dt_object.strftime("%d-%m-%y %H:%M")
    return formatted_date

def get_new_api_key(current_index) -> int:
    """
        Gets a new api key form a list of key, returns the index of the api key
    """
    with open("api_keys.json", 'r') as f:
        keys_json:list[str] = json.load(fp=f)

    current_index+=1

    try:
        new_key=keys_json[current_index]
        
    except IndexError:
        raise Exception("No valid keys found")
    
    print("writing new key starting to env file", new_key[:4])

    os.environ['ODDS_API_KEY']=new_key
    set_key(".env","ODDS_API_KEY", new_key)

    return current_index
    
    
